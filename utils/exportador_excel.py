"""exportador_excel.py - Sistema de exportação Excel com fallback robusto"""

import pandas as pd
from pathlib import Path
from datetime import datetime
import io

# Verificação segura de bibliotecas opcionais
OPENPYXL_AVAILABLE = False
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    # Bibliotecas não disponíveis - usaremos fallback
    pass

# Cores Amaro Aviation (em hex para Excel)
AMARO_PRIMARY = "8C1D40"
AMARO_SECONDARY = "A02050"  
AMARO_LIGHT = "F8F9FA"
AMARO_SUCCESS = "27AE60"

def gerar_excel_simples(buffer_arquivo, dados: dict):
    """
    Gera planilha Excel simples usando apenas pandas
    """
    try:
        # Preparar dados estruturados
        relatorio_data = []
        
        # Cabeçalho
        relatorio_data.extend([
            ["", ""],
            ["✈️ AMARO AVIATION", ""],
            ["RELATÓRIO DE ANÁLISE DE CUSTOS", ""],
            [f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ""],
            ["", ""],
            ["INFORMAÇÕES GERAIS", ""],
        ])
        
        # Informações básicas
        informacoes = [
            ("Tipo de Análise:", dados.get('Análise', 'Não especificado')),
            ("Modelo da Aeronave:", dados.get('Modelo', 'Não especificado')),
            ("Rota:", dados.get('Rota', 'Não especificado')),
            ("Duração do Voo:", dados.get('Duração', 'Não especificado')),
        ]
        
        for campo, valor in informacoes:
            relatorio_data.append([campo, str(valor)])
        
        relatorio_data.append(["", ""])
        relatorio_data.append(["BREAKDOWN DE CUSTOS", ""])
        
        # Custos detalhados
        componentes = [
            ('Combustível', dados.get('Combustível', 0)),
            ('Piloto', dados.get('Piloto', 0)),
            ('Manutenção', dados.get('Manutenção', 0)),
            ('Depreciação', dados.get('Depreciação', 0))
        ]
        
        total_custos = 0
        for nome, valor in componentes:
            if isinstance(valor, (int, float)):
                total_custos += valor
                relatorio_data.append([nome, f"R$ {valor:,.2f}"])
        
        relatorio_data.append(["TOTAL", f"R$ {total_custos:,.2f}"])
        relatorio_data.append(["", ""])
        
        # Comparativo
        relatorio_data.append(["COMPARATIVO COM MERCADO", ""])
        custo_amaro = dados.get('Custo Total Amaro', total_custos)
        preco_mercado = dados.get('Preço Mercado', 0)
        economia = dados.get('Economia', 0)
        
        relatorio_data.extend([
            ["Custo Amaro Aviation", f"R$ {custo_amaro:,.2f}"],
            ["Preço de Mercado", f"R$ {preco_mercado:,.2f}"],
            ["Economia", f"R$ {economia:,.2f}"],
        ])
        
        if preco_mercado > 0:
            percentual = (economia / preco_mercado) * 100
            relatorio_data.append(["Percentual de Economia", f"{percentual:.1f}%"])
        
        # Converter para DataFrame
        df = pd.DataFrame(relatorio_data, columns=["Campo", "Valor"])
        
        # Salvar arquivo
        if isinstance(buffer_arquivo, (str, Path)):
            df.to_excel(buffer_arquivo, index=False, sheet_name="Relatório Amaro")
        else:
            df.to_excel(buffer_arquivo, index=False, sheet_name="Relatório Amaro")
        
        return True
        
    except Exception as e:
        print(f"Erro ao gerar Excel simples: {e}")
        return False

def criar_estilos_openpyxl():
    """Cria estilos para openpyxl (apenas se disponível)"""
    if not OPENPYXL_AVAILABLE:
        return {}
    
    estilos = {}
    
    # Fontes
    estilos['header_font'] = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    estilos['title_font'] = Font(name='Calibri', size=18, bold=True, color=AMARO_PRIMARY)
    estilos['normal_font'] = Font(name='Calibri', size=11, color='000000')
    estilos['highlight_font'] = Font(name='Calibri', size=11, bold=True, color=AMARO_PRIMARY)
    
    # Preenchimentos
    estilos['header_fill'] = PatternFill(start_color=AMARO_PRIMARY, end_color=AMARO_PRIMARY, fill_type='solid')
    estilos['secondary_fill'] = PatternFill(start_color=AMARO_SECONDARY, end_color=AMARO_SECONDARY, fill_type='solid')
    estilos['light_fill'] = PatternFill(start_color=AMARO_LIGHT, end_color=AMARO_LIGHT, fill_type='solid')
    estilos['success_fill'] = PatternFill(start_color=AMARO_SUCCESS, end_color=AMARO_SUCCESS, fill_type='solid')
    
    # Alinhamentos
    estilos['center_align'] = Alignment(horizontal='center', vertical='center', wrap_text=True)
    estilos['left_align'] = Alignment(horizontal='left', vertical='center', wrap_text=True)
    estilos['right_align'] = Alignment(horizontal='right', vertical='center')
    
    # Bordas
    thin_border = Side(border_style='thin', color='000000')
    estilos['border'] = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    return estilos

def gerar_excel_premium_openpyxl(buffer_arquivo, dados: dict):
    """
    Gera Excel premium usando openpyxl (apenas se disponível)
    """
    if not OPENPYXL_AVAILABLE:
        return gerar_excel_simples(buffer_arquivo, dados)
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Análise de Custos"
        
        estilos = criar_estilos_openpyxl()
        
        # Cabeçalho
        ws['A1'] = '✈️ AMARO AVIATION'
        if estilos:
            ws['A1'].font = estilos['title_font']
            ws['A1'].alignment = estilos['center_align']
        ws.merge_cells('A1:F1')
        
        ws['A2'] = 'RELATÓRIO DE ANÁLISE DE CUSTOS'
        if estilos:
            ws['A2'].font = estilos['highlight_font']
            ws['A2'].alignment = estilos['center_align']
        ws.merge_cells('A2:F2')
        
        ws['A3'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")}'
        if estilos:
            ws['A3'].font = estilos['normal_font']
            ws['A3'].alignment = estilos['center_align']
        ws.merge_cells('A3:F3')
        
        linha = 5
        
        # Seção de informações
        ws[f'A{linha}'] = 'INFORMAÇÕES GERAIS'
        if estilos:
            ws[f'A{linha}'].font = estilos['header_font']
            ws[f'A{linha}'].fill = estilos['header_fill']
            ws[f'A{linha}'].alignment = estilos['center_align']
        ws.merge_cells(f'A{linha}:B{linha}')
        linha += 1
        
        informacoes = [
            ('Tipo de Análise:', dados.get('Análise', 'Não especificado')),
            ('Modelo da Aeronave:', dados.get('Modelo', 'Não especificado')),
            ('Rota:', dados.get('Rota', 'Não especificado')),
            ('Duração do Voo:', dados.get('Duração', 'Não especificado')),
        ]
        
        for campo, valor in informacoes:
            ws[f'A{linha}'] = campo
            ws[f'B{linha}'] = str(valor)
            if estilos:
                ws[f'A{linha}'].font = estilos['highlight_font']
                ws[f'B{linha}'].font = estilos['normal_font']
                ws[f'A{linha}'].border = estilos['border']
                ws[f'B{linha}'].border = estilos['border']
            linha += 1
        
        linha += 1
        
        # Seção de custos
        ws[f'A{linha}'] = 'BREAKDOWN DE CUSTOS'
        if estilos:
            ws[f'A{linha}'].font = estilos['header_font']
            ws[f'A{linha}'].fill = estilos['header_fill']
            ws[f'A{linha}'].alignment = estilos['center_align']
        ws.merge_cells(f'A{linha}:C{linha}')
        linha += 1
        
        # Cabeçalhos da tabela de custos
        headers = ['Componente', 'Valor (R$)', 'Percentual']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=linha, column=col, value=header)
            if estilos:
                cell.font = estilos['header_font']
                cell.fill = estilos['secondary_fill']
                cell.alignment = estilos['center_align']
                cell.border = estilos['border']
        linha += 1
        
        # Dados dos custos
        componentes = [
            ('Combustível', dados.get('Combustível', 0)),
            ('Piloto', dados.get('Piloto', 0)),
            ('Manutenção', dados.get('Manutenção', 0)),
            ('Depreciação', dados.get('Depreciação', 0))
        ]
        
        total_custos = sum(valor for _, valor in componentes if isinstance(valor, (int, float)))
        
        for nome, valor in componentes:
            if isinstance(valor, (int, float)):
                percentual = (valor / total_custos * 100) if total_custos > 0 else 0
                
                ws[f'A{linha}'] = nome
                ws[f'B{linha}'] = valor
                ws[f'C{linha}'] = percentual / 100
                
                if estilos:
                    ws[f'A{linha}'].font = estilos['normal_font']
                    ws[f'B{linha}'].font = estilos['normal_font']
                    ws[f'C{linha}'].font = estilos['normal_font']
                    ws[f'A{linha}'].border = estilos['border']
                    ws[f'B{linha}'].border = estilos['border']
                    ws[f'C{linha}'].border = estilos['border']
                    ws[f'B{linha}'].number_format = 'R$ #,##0.00'
                    ws[f'C{linha}'].number_format = '0.0%'
                
                linha += 1
        
        # Ajustar larguras das colunas
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20
        
        # Adicionar aba de dados brutos
        ws_dados = wb.create_sheet("Dados Brutos")
        dados_df = pd.DataFrame(list(dados.items()), columns=['Campo', 'Valor'])
        
        for r in dataframe_to_rows(dados_df, index=False, header=True):
            ws_dados.append(r)
        
        # Salvar arquivo
        if isinstance(buffer_arquivo, (str, Path)):
            wb.save(buffer_arquivo)
        else:
            wb.save(buffer_arquivo)
        
        return True
        
    except Exception as e:
        print(f"Erro ao gerar Excel premium: {e}")
        return gerar_excel_simples(buffer_arquivo, dados)

def gerar_excel(buffer_arquivo, dados: dict):
    """
    Função principal para gerar Excel com fallback automático
    
    Args:
        buffer_arquivo: Buffer para salvar (BytesIO ou caminho do arquivo)
        dados: Dicionário com os dados para o relatório
        
    Returns:
        bool: True se sucesso, False se erro
    """
    try:
        # Validar dados de entrada
        if not dados or not isinstance(dados, dict):
            print("Dados inválidos para gerar Excel")
            return False
        
        # Tentar versão premium se openpyxl estiver disponível
        if OPENPYXL_AVAILABLE:
            return gerar_excel_premium_openpyxl(buffer_arquivo, dados)
        else:
            # Fallback para versão simples com pandas
            return gerar_excel_simples(buffer_arquivo, dados)
            
    except Exception as e:
        print(f"Erro geral ao gerar Excel: {e}")
        
        # Último recurso - gerar CSV
        try:
            df = pd.DataFrame(list(dados.items()), columns=["Campo", "Valor"])
            
            if isinstance(buffer_arquivo, (str, Path)):
                # Trocar extensão para CSV
                arquivo_csv = str(buffer_arquivo).replace('.xlsx', '.csv').replace('.xls', '.csv')
                df.to_csv(arquivo_csv, index=False)
                print(f"Arquivo salvo como CSV: {arquivo_csv}")
            else:
                df.to_csv(buffer_arquivo, index=False)
                print("Dados salvos como CSV no buffer")
            
            return True
            
        except Exception as csv_error:
            print(f"Erro ao gerar CSV como fallback: {csv_error}")
            return False

# Função de compatibilidade (caso seja chamada diretamente)
def gerar_excel_compatibilidade(caminho_arquivo, dados: dict):
    """
    Função de compatibilidade para manter interface anterior
    """
    return gerar_excel(caminho_arquivo, dados)